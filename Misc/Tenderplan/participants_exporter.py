from Templates.make_excel import make2exel
import Misc.Tenderplan.participants2excel as p2ex

import openpyxl

excel_name = r'E:\Лиды_экспорт.xlsx'

make2exel(
    ['Название компании', 'ИНН', 'Список выигранных тендеров', 'Список остальных тендеров c участием'], excel_name)

wb = openpyxl.load_workbook(excel_name)
sheet = wb.active

participants_dict_list = p2ex.get_participants()


curr_row = 1
for x in range(len(participants_dict_list)):
    curr_row += 1
    print(participants_dict_list[x]['company'])
    sheet.cell(row=curr_row, column=1).value = participants_dict_list[x]['company']

    for t in range(len(participants_dict_list[x]['tender_name'])):
        sheet.cell(row=curr_row, column=4).value = participants_dict_list[x]['tender_name'][t]
        curr_row += 1
        # if t > len(participants_dict_list[x]['tender_name']):
        #     curr_row += 1
wb.save(excel_name)

