import openpyxl
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font

from SCANEX.zip_meta2exel.shp import scanex_archives_export


def metadata2exel(zip_dir, out_filename):
    """
    На входе берётся директория со скачанными с search.kosmosnimki.ru архивами покрытия (zip_dir)
    На выходе - excel-файл (out_filename) с выбранными метаданными
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Main'
    arial_font = Font(name='Arial', bold=True, size=11)
    column_names_list = ['VENDOR_ID', 'DATE', 'SAT_NAME', 'OFF_NADIR', 'SUN_ELEV', 'CLOUD_PCT']
    for i in range(len(column_names_list)):
        sheet.cell(row=1, column=i + 1).value = column_names_list[i]
        sheet.cell(row=1, column=i + 1).font = arial_font
        sheet.column_dimensions[get_column_letter(i + 1)].width = 38.5
    # закрепляем шапку
    sheet.freeze_panes = 'A2'
    # импортируем список найденнных снимков с метаданными
    img_list = scanex_archives_export(zip_dir)
    for x in range(len(img_list)):
        sheet.cell(row=x+2, column=1).value = img_list[x]['vendor_id']
        sheet.cell(row=x+2, column=2).value = img_list[x]['date']
        sheet.cell(row=x+2, column=3).value = img_list[x]['sat_name']
        sheet.cell(row=x+2, column=4).value = img_list[x]['off_nadir']
        sheet.cell(row=x+2, column=5).value = img_list[x]['sun_elev']
        sheet.cell(row=x+2, column=6).value = img_list[x]['cloud_pct']

    wb.save(out_filename)
