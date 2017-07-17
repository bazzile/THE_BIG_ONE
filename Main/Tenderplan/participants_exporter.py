from Templates.make_excel import make2exel
import Main.Tenderplan.participants2excel as p2ex

import openpyxl

excel_name = r'E:\Лиды_экспорт.xlsx'

make2exel(
    ['Название компании', 'ИНН', 'Список выигранных тендеров', 'Список остальных тендеров c участием'], excel_name)

wb = openpyxl.load_workbook(excel_name)
sheet = wb.active

participants_dict_list = p2ex.get_participants()


curr_row = 1
for x in range(len(participants_dict_list)):
    curr_row += 1
    print(participants_dict_list[x]['company'])
    sheet.cell(row=curr_row, column=1).value = participants_dict_list[x]['company']

    part_row = win_row = curr_row
    for t in range(len(participants_dict_list[x]['tender_name'])):
        if list(participants_dict_list[x]['tender_name'][t].values())[0] == 'par':
            sheet.cell(row=part_row, column=4).value = (list(participants_dict_list[x]['tender_name'][t].keys())[0])
            part_row += 1
        else:
            sheet.cell(row=win_row, column=3).value = (list(participants_dict_list[x]['tender_name'][t].keys())[0])
            win_row += 1
        curr_row += 1
        # if t > len(participants_dict_list[x]['tender_name']):
        #     curr_row += 1
wb.save(excel_name)

